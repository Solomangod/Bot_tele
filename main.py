
import os
import asyncio
import threading
from flask import Flask, request
from telegram import Update, InputFile
from telegram.ext import Application, CommandHandler, MessageHandler, ContextTypes, filters
from openpyxl import load_workbook
from io import BytesIO
from dotenv import load_dotenv

load_dotenv()
TOKEN = os.getenv("BOT_TOKEN")
WEBHOOK_SECRET = os.getenv("WEBHOOK_SECRET", "supersecret")
BASE_URL = os.getenv("BASE_URL")

if not TOKEN or not BASE_URL:
    raise RuntimeError("BOT_TOKEN v√† BASE_URL c·∫ßn ƒë∆∞·ª£c khai b√°o trong .env")

app = Flask(__name__)
telegram_app = Application.builder().token(TOKEN).build()
user_data = {}
stop_flags = {}

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("üì• G·ª≠i file .txt ch·ª©a t·ª´ kh√≥a tr∆∞·ªõc, sau ƒë√≥ g·ª≠i file Excel .xlsx.\nG√µ /stop ƒë·ªÉ d·ª´ng qu√° tr√¨nh x·ª≠ l√Ω.")

async def stop(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.message.chat_id
    stop_flags[chat_id] = True
    await update.message.reply_text("‚èπ ƒê√£ g·ª≠i y√™u c·∫ßu d·ª´ng.")

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message or not update.message.document:
        return
    doc = update.message.document
    file_name = doc.file_name.lower()
    if file_name.endswith(".txt"):
        await handle_txt(update, context)
    elif file_name.endswith(".xlsx"):
        await handle_excel(update, context)
    else:
        await update.message.reply_text("‚ö†Ô∏è Ch·ªâ h·ªó tr·ª£ file .txt v√† .xlsx")

async def handle_txt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    file = await update.message.document.get_file()
    content = await file.download_as_bytearray()
    keywords = set(kw.strip().lower() for kw in content.decode("utf-8").splitlines() if kw.strip())
    user_data[update.message.chat_id] = {"keywords": keywords}
    await update.message.reply_text(f"‚úÖ ƒê√£ nh·∫≠n {len(keywords)} t·ª´ kh√≥a. G·ª≠i file Excel ti·∫øp theo.")

async def handle_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.message.chat_id
    stop_flags[chat_id] = False
    if chat_id not in user_data:
        await update.message.reply_text("‚ö†Ô∏è B·∫°n ch∆∞a g·ª≠i file .txt ch·ª©a t·ª´ kh√≥a.")
        return
    file = await update.message.document.get_file()
    xlsx_bytes = await file.download_as_bytearray()
    wb = load_workbook(filename=BytesIO(xlsx_bytes))
    ws = wb.active
    keywords = user_data[chat_id]["keywords"]
    rows = [row for row in ws.iter_rows(min_row=2, max_col=1) if row[0].value]
    total = len(rows)
    if total == 0:
        await update.message.reply_text("‚ùå File Excel kh√¥ng c√≥ d·ªØ li·ªáu ·ªü c·ªôt A.")
        return
    match_count = 0
    progress_message = await update.message.reply_text("üîÑ B·∫Øt ƒë·∫ßu x·ª≠ l√Ω file...")
    for idx, row in enumerate(rows, start=1):
        if stop_flags.get(chat_id):
            await update.message.reply_text("‚èπ ƒê√£ d·ª´ng theo y√™u c·∫ßu.")
            return
        cell = row[0].value
        text = str(cell).lower() if cell else ""
        words = set(text.replace(",", " ").replace(".", " ").replace("!", " ").replace("?", " ").split())
        found = any(kw in words for kw in keywords)
        row[0].offset(column=6).value = "SOS: N√≥ kia k√¨a √êTH" if found else ""
        if found:
            match_count += 1
        if idx % max(1, total // 100) == 0 or idx == total:
            percent = int((idx / total) * 100)
            await progress_message.edit_text(f"üîÑ {percent}% ({idx}/{total} d√≤ng)")
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    await progress_message.edit_text(f"‚úÖ Xong! {match_count}/{total} Ti√™u ƒë·ªÅ c√≥ t·ª´ vi ph·∫°m.")
    await update.message.reply_document(document=InputFile(output, filename="Checked_Results.xlsx"))
    user_data.pop(chat_id, None)

loop = asyncio.new_event_loop()
threading.Thread(target=loop.run_forever, daemon=True).start()

telegram_app.add_handler(CommandHandler("start", start))
loop = asyncio.new_event_loop()
threading.Thread(target=loop.run_forever, daemon=True).start()

telegram_app.add_handler(CommandHandler("stop", stop))
loop = asyncio.new_event_loop()
threading.Thread(target=loop.run_forever, daemon=True).start()

telegram_app.add_handler(MessageHandler(filters.Document.ALL, handle_document))

@app.route("/")
def index():
    return "‚úÖ Bot Telegram ƒëang ch·∫°y b·∫±ng webhook tr√™n Flask!"



@app.route(f"/webhook/{WEBHOOK_SECRET}", methods=["POST"])
def telegram_webhook():
    update = Update.de_json(request.get_json(force=True), telegram_app.bot)

    async def run_update():
        await telegram_app.process_update(update)

    future = asyncio.run_coroutine_threadsafe(run_update(), loop)
    try:
        future.result()
    except Exception as e:
        print("‚ùå Error in coroutine:", e)
        return "Error", 500

    return "OK"



async def set_webhook():
    url = f"{BASE_URL}/webhook/{WEBHOOK_SECRET}"
    await telegram_app.bot.set_webhook(url)
    print("üåê ƒê√£ ƒëƒÉng k√Ω Webhook:", url)

if __name__ == "__main__":
    asyncio.run(set_webhook())
    asyncio.run(telegram_app.initialize())
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
